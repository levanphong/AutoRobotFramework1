import logging

from openpyxl import load_workbook

from src.models.StepTest import StepTest
from src.models.TestcaseInfo import TestcaseInfo
from src.models.TestFeature import TestFeature


class ExcelLibReader(object):
    def __init__(self):
        logging.log(logging.INFO, "ExcelLibReader")

    def read_sheet(self, workbook, sheet_name):
        sh = workbook[sheet_name]
        steps = []
        testcases = []
        current_name = ""
        current_id = ""
        current_precondition = []
        # iterate through excel and display data
        for i in range(1, sh.max_row + 1):
            print("Row ", i, " data :", "\n")
            if i == 1:
                continue
            id_test = sh.cell(i, 1).value
            name = sh.cell(i, 2).value
            precondition = sh.cell(i, 3).value
            if id_test is not None:
                if steps:
                    testcase_item = TestcaseInfo()
                    testcase_item.set_name(current_name)
                    testcase_item.set_id(current_id)
                    testcase_item.set_precondition(current_precondition)
                    testcase_item.set_steps(steps)
                    testcases.append(testcase_item)
                    steps = []
                if id_test == "end":
                    break
                current_id = id_test
                current_name = name
                current_precondition = precondition
            step_name = sh.cell(i, 4).value
            step_data_test = sh.cell(i, 5).value
            step_expected_result = sh.cell(i, 6).value
            step_item = StepTest(step_name, step_data_test, step_expected_result)
            steps.append(step_item)
        return testcases

    def read_all_sheet(self, file_path):
        workbook = load_workbook(file_path)
        test_features = []
        for sheet in workbook.sheetnames:
            test_suite = TestFeature()
            test_suite.set_name(sheet)
            test_suite.set_testCase_list(self.read_sheet(workbook, sheet))
            test_features.append(test_suite)
        print(len(test_features))
        return test_features
